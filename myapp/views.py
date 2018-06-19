from django.shortcuts import render
import openpyxl
import requests
import csv
from django.conf import settings
from django.http import HttpResponse

def index(request):
    static = settings.STATIC_URL
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {"static":static})
    else:
        if request.FILES.has_key("excel_file"):
            excel_file = request.FILES["excel_file"]
            ext = request.FILES['excel_file'].name.split('.')[-1].lower()
            if ext == "xls" or ext == "xlsx":
                wb = openpyxl.load_workbook(excel_file)
                sheets = wb.sheetnames
                worksheet = wb["Sheet1"]
                active_sheet = wb.active
                geocode_result = list()
                excel_data = list()
                error = ''

                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        if cell.coordinate=='A1' and cell.value!='Address':
                            error = 'Uploaded excel format is not correct. Please dowload sample upload and try again'
                            break
                        if cell.value and cell.coordinate!='A1':
                            try:
                                geocode_result = get_google_results(cell.value, settings.GOOGLE_API_KEY)
                            except Exception as e:
                                print(e)
                                error = 'Oops! Some Server Error Occurred'
                                break
                            if geocode_result['status'] == 'OVER_QUERY_LIMIT':
                                error = 'Your Dialy Quota Over'
                                break
                            else:
                                if geocode_result['status'] == 'OK':
                                    row_data.append(geocode_result['input_string'])
                                    row_data.append(geocode_result['formatted_address'])
                                    row_data.append(geocode_result['latitude'])
                                    row_data.append(geocode_result['longitude'])
                                else:
                                    row_data.append(cell.value)
                                    row_data.append('NA')
                                    row_data.append('0')
                                    row_data.append('0')
                    if len(error):
                        break
                    if(len(row_data)):
                        excel_data.append(row_data)
                if len(error)==0:
                    if len(excel_data):
                        response = HttpResponse(content_type='text/csv')
                        response['Content-Disposition'] = 'attachment; filename="addresses-latitude-longitude.csv"'
                        writer = csv.writer(response)
                        writer.writerow(['Address','Formatted Address', 'Latitude', 'Longitude'])
                        for r in excel_data:
                            writer.writerow(r)
                        return response;
                    else:
                        error = 'Uploaded Excel Sheet is Empty. Please upload correct file and try again'
            else:
                error = 'File format should be XLS or XLSX'
        else:
            error = 'Please upload file to convert'
        return render(request, 'myapp/index.html', {"error":error,"static":static})


#------------------ Get latitude and longitude using google geo code ------------------------

def get_google_results(address, api_key=None):
    geocode_url = "https://maps.googleapis.com/maps/api/geocode/json?address={}".format(address)
    if api_key is not None:
        geocode_url = geocode_url + "&key={}".format(api_key)
        
    results = requests.get(geocode_url)
    results = results.json()
    
    if len(results['results']) == 0:
        output = {
            "formatted_address" : None,
            "latitude": None,
            "longitude": None,
            "accuracy": None,
            "google_place_id": None,
            "type": None,
            "postcode": None
        }
    else:    
        answer = results['results'][0]
        output = {
            "formatted_address" : answer.get('formatted_address'),
            "latitude": answer.get('geometry').get('location').get('lat'),
            "longitude": answer.get('geometry').get('location').get('lng'),
            "accuracy": answer.get('geometry').get('location_type'),
            "google_place_id": answer.get("place_id"),
            "type": ",".join(answer.get('types')),
            "postcode": ",".join([x['long_name'] for x in answer.get('address_components') 
                                  if 'postal_code' in x.get('types')])
        }
        
    # Append some other details:    
    output['input_string'] = address
    output['number_of_results'] = len(results['results'])
    output['status'] = results.get('status')
    output['response'] = results
    
    return output